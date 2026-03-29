# -*- coding: utf-8 -*-
"""
===============================================================================
Title:       EEG Decoding: Transformer vs. RNN vs. LSTM vs. EEGNet vs. HeteroRC
Author:      Based on HeteroRC project by Runhao Lu
Date:        Mar 2026
Description:
    Benchmarks Standard and Windowed variants (hidden=32) of Transformer, RNN,
    LSTM, plus EEGNet, against HeteroRC for time-resolved EEG decoding.

    Left plot : Standard RNN, LSTM, Transformer (32) + HeteroRC + EEGNet
    Right plot: Windowed RNN, LSTM, Transformer (32) + HeteroRC + EEGNet

    N=30 subjects.  Per-subject accuracy and per-model timing saved to Excel.
===============================================================================
"""

import math
import os
import time as time_module
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from scipy.stats import sem
from scipy.ndimage import gaussian_filter1d
from sklearn.model_selection import StratifiedKFold
import torch
import torch.nn as nn
from torch.utils.data import Dataset, DataLoader, TensorDataset

from heterorc import time_resolved_decoding_heterorc
from simulate_eeg import simulate_data

# ========================= 1. Global Configuration =========================
N_SUBJECTS = 30
SFREQ = 100
TMIN, TMAX = 0.0, 0.8
N_CHANNELS = 32
N_TOTAL_TRIALS = 80

SIM_MODE = "erp"
TARGET_FREQ = 10
N_FOLDS = 5
ROBUST_PCTL = 99

# --- Model specifics ---
HIDDEN_SIZE = 32
DROPOUT = 0.5
LR = 0.002
EPOCHS = 30
BATCH_SIZE = 64
DEVICE = "cuda" if torch.cuda.is_available() else "cpu"
WINDOW_SIZE = 10  # 100 ms

# --- Transformer specifics ---
N_TRANSFORMER_LAYERS = 2
DIM_FEEDFORWARD_MULT = 4

# --- EEGNet specifics ---
EEGNET_LR = 0.005
EEGNET_EPOCHS = 25

# --- HeteroRC specifics ---
rc_params = dict(
    n_res=350, input_scaling=0.5, bias_scaling=0.5, spectral_radius=0.95,
    tau_mode=0.01, tau_sigma=0.8, bidirectional=True, merge_mode="product", fs=SFREQ,
)

print(f"Running on {DEVICE}. Hidden={HIDDEN_SIZE}, Dropout={DROPOUT}, Window={WINDOW_SIZE}")


# ========================= 2. Data Helpers =========================

class EEGDataset(Dataset):
    """Wraps (n_trials, n_channels, n_times) into (n_trials, n_times, n_channels)."""
    def __init__(self, X, y):
        self.X = torch.tensor(X.transpose(0, 2, 1), dtype=torch.float32)
        self.y = torch.tensor(y, dtype=torch.long)

    def __len__(self):
        return self.X.shape[0]

    def __getitem__(self, idx):
        return self.X[idx], self.y[idx]


def make_sliding_windows(X, window_size):
    """Causal sliding windows for RNN/LSTM/Transformer windowed models."""
    n_trials, n_ch, n_times = X.shape
    X_pad = np.pad(X, ((0, 0), (0, 0), (window_size - 1, 0)), mode='constant')
    X_windows = []
    for t in range(n_times):
        win = X_pad[:, :, t: t + window_size]
        X_windows.append(win)
    X_windows = np.stack(X_windows, axis=1).transpose(0, 1, 3, 2)
    return X_windows.reshape(-1, window_size, n_ch)


def make_sliding_windows_eegnet(X, window_size):
    """Causal sliding windows for EEGNet (4D: batch, 1, channels, time)."""
    n_trials, n_ch, n_times = X.shape
    X_pad = np.pad(X, ((0, 0), (0, 0), (window_size - 1, 0)), mode='constant')
    X_windows = []
    for t in range(n_times):
        win = X_pad[:, :, t: t + window_size]
        X_windows.append(win)
    X_windows = np.stack(X_windows, axis=1)
    X_flat = X_windows.reshape(-1, n_ch, window_size)
    return X_flat[:, np.newaxis, :, :]


# ========================= 3. Positional Encoding =========================

class SinusoidalPositionalEncoding(nn.Module):
    def __init__(self, d_model, max_len=500, dropout=0.1):
        super().__init__()
        self.dropout = nn.Dropout(p=dropout)
        pe = torch.zeros(max_len, d_model)
        position = torch.arange(0, max_len, dtype=torch.float).unsqueeze(1)
        div_term = torch.exp(
            torch.arange(0, d_model, 2).float() * (-math.log(10000.0) / d_model)
        )
        pe[:, 0::2] = torch.sin(position * div_term)
        if d_model % 2 == 1:
            pe[:, 1::2] = torch.cos(position * div_term[:-1])
        else:
            pe[:, 1::2] = torch.cos(position * div_term)
        pe = pe.unsqueeze(0)
        self.register_buffer('pe', pe)

    def forward(self, x):
        x = x + self.pe[:, :x.size(1), :]
        return self.dropout(x)


# ========================= 4. Model Definitions =========================

# --------------- Standard (Seq-to-Seq) ---------------

class StandardRNN(nn.Module):
    def __init__(self, n_in, n_hidden, n_classes, dropout=0.0):
        super().__init__()
        self.rnn = nn.RNN(n_in, n_hidden, batch_first=True, nonlinearity='tanh')
        self.drop = nn.Dropout(dropout)
        self.head = nn.Linear(n_hidden, n_classes)

    def forward(self, x):
        out, _ = self.rnn(x)
        return self.head(self.drop(out))


class StandardLSTM(nn.Module):
    def __init__(self, n_in, n_hidden, n_classes, dropout=0.0):
        super().__init__()
        self.lstm = nn.LSTM(n_in, n_hidden, batch_first=True)
        self.drop = nn.Dropout(dropout)
        self.head = nn.Linear(n_hidden, n_classes)

    def forward(self, x):
        out, _ = self.lstm(x)
        return self.head(self.drop(out))


class StandardTransformer(nn.Module):
    def __init__(self, n_in, n_hidden, n_classes, dropout=0.0):
        super().__init__()
        d_model = n_hidden
        n_heads = max(1, d_model // 8)
        self.input_proj = nn.Linear(n_in, d_model)
        self.pos_enc = SinusoidalPositionalEncoding(d_model, max_len=500, dropout=dropout)
        encoder_layer = nn.TransformerEncoderLayer(
            d_model=d_model, nhead=n_heads,
            dim_feedforward=d_model * DIM_FEEDFORWARD_MULT,
            dropout=dropout, batch_first=True,
        )
        self.transformer = nn.TransformerEncoder(encoder_layer, num_layers=N_TRANSFORMER_LAYERS)
        self.head = nn.Linear(d_model, n_classes)

    def forward(self, x):
        x = self.pos_enc(self.input_proj(x))
        return self.head(self.transformer(x))


# --------------- Windowed (Seq-to-One) ---------------

class WindowedRNN(nn.Module):
    def __init__(self, n_in, n_hidden, n_classes, dropout=0.0):
        super().__init__()
        self.rnn = nn.RNN(n_in, n_hidden, batch_first=True, nonlinearity='tanh')
        self.drop = nn.Dropout(dropout)
        self.head = nn.Linear(n_hidden, n_classes)

    def forward(self, x):
        _, h_n = self.rnn(x)
        return self.head(self.drop(h_n[-1]))


class WindowedLSTM(nn.Module):
    def __init__(self, n_in, n_hidden, n_classes, dropout=0.0):
        super().__init__()
        self.lstm = nn.LSTM(n_in, n_hidden, batch_first=True)
        self.drop = nn.Dropout(dropout)
        self.head = nn.Linear(n_hidden, n_classes)

    def forward(self, x):
        _, (h_n, _) = self.lstm(x)
        return self.head(self.drop(h_n[-1]))


class WindowedTransformer(nn.Module):
    def __init__(self, n_in, n_hidden, n_classes, dropout=0.0):
        super().__init__()
        d_model = n_hidden
        n_heads = max(1, d_model // 8)
        self.input_proj = nn.Linear(n_in, d_model)
        self.pos_enc = SinusoidalPositionalEncoding(d_model, max_len=500, dropout=dropout)
        encoder_layer = nn.TransformerEncoderLayer(
            d_model=d_model, nhead=n_heads,
            dim_feedforward=d_model * DIM_FEEDFORWARD_MULT,
            dropout=dropout, batch_first=True,
        )
        self.transformer = nn.TransformerEncoder(encoder_layer, num_layers=N_TRANSFORMER_LAYERS)
        self.drop = nn.Dropout(dropout)
        self.head = nn.Linear(d_model, n_classes)

    def forward(self, x):
        x = self.pos_enc(self.input_proj(x))
        x = self.transformer(x).mean(dim=1)
        return self.head(self.drop(x))


# --------------- EEGNet ---------------

class EEGNet_Small(nn.Module):
    def __init__(self, n_channels, n_classes, window_size=10, F1=8, D=2, F2=16):
        super().__init__()
        self.conv1 = nn.Conv2d(1, F1, (1, window_size // 2),
                               padding=(0, window_size // 4), bias=False)
        self.bn1 = nn.BatchNorm2d(F1)
        self.conv2 = nn.Conv2d(F1, F1 * D, (n_channels, 1), groups=F1, bias=False)
        self.bn2 = nn.BatchNorm2d(F1 * D)
        self.elu1 = nn.ELU()
        self.pool1 = nn.AvgPool2d((1, 4))
        self.drop1 = nn.Dropout(0.25)
        self.conv3 = nn.Conv2d(F1 * D, F2, (1, 8), padding=(0, 4), groups=F1 * D, bias=False)
        self.conv4 = nn.Conv2d(F2, F2, (1, 1), bias=False)
        self.bn3 = nn.BatchNorm2d(F2)
        self.elu2 = nn.ELU()
        self.pool2 = nn.AvgPool2d((1, 2))
        self.drop2 = nn.Dropout(0.25)
        with torch.no_grad():
            dummy = torch.zeros(1, 1, n_channels, window_size)
            z = self._features(dummy)
            self.flat_dim = z.view(1, -1).shape[1]
        self.classifier = nn.Linear(self.flat_dim, n_classes)

    def _features(self, x):
        x = self.bn1(self.conv1(x))
        x = self.elu1(self.bn2(self.conv2(x)))
        x = self.drop1(self.pool1(x))
        x = self.bn3(self.conv4(self.conv3(x)))
        x = self.elu2(x)
        if x.shape[3] > 1:
            x = self.pool2(x)
        return self.drop2(x)

    def forward(self, x):
        x = self._features(x)
        return self.classifier(x.view(x.size(0), -1))


# ========================= 5. Training Loops =========================

def train_standard_model(model_class, X_train, y_train, X_test, y_test):
    """Train a standard (seq-to-seq) model and return per-time-point accuracy."""
    ds_train = EEGDataset(X_train, y_train)
    ds_test = EEGDataset(X_test, y_test)
    dl_train = DataLoader(ds_train, batch_size=BATCH_SIZE, shuffle=True)
    dl_test = DataLoader(ds_test, batch_size=BATCH_SIZE, shuffle=False)

    model = model_class(N_CHANNELS, HIDDEN_SIZE, 2, dropout=DROPOUT).to(DEVICE)
    opt = torch.optim.Adam(model.parameters(), lr=LR, weight_decay=1e-4)
    crit = nn.CrossEntropyLoss()

    model.train()
    for _ in range(EPOCHS):
        for xb, yb in dl_train:
            xb, yb = xb.to(DEVICE), yb.to(DEVICE)
            logits = model(xb)
            y_rep = yb[:, None].expand(-1, logits.shape[1])
            loss = crit(logits.reshape(-1, 2), y_rep.reshape(-1))
            opt.zero_grad(); loss.backward(); opt.step()

    model.eval()
    all_logits = []
    with torch.no_grad():
        for xb, _ in dl_test:
            all_logits.append(model(xb.to(DEVICE)).cpu().numpy())
    logits = np.concatenate(all_logits, axis=0)
    preds = np.argmax(logits, axis=-1)
    return np.mean(preds == y_test[:, None], axis=0)


def train_windowed_model(model_class, X_train, y_train, X_test, y_test):
    """Train a windowed (seq-to-one) model and return per-time-point accuracy."""
    X_tr_w = make_sliding_windows(X_train, WINDOW_SIZE)
    X_te_w = make_sliding_windows(X_test, WINDOW_SIZE)
    n_times = X_train.shape[2]
    y_tr_flat = np.repeat(y_train[:, None], n_times, axis=1).flatten()

    ds_train = TensorDataset(
        torch.tensor(X_tr_w, dtype=torch.float32),
        torch.tensor(y_tr_flat, dtype=torch.long),
    )
    dl_train = DataLoader(ds_train, batch_size=BATCH_SIZE * 2, shuffle=True)

    model = model_class(N_CHANNELS, HIDDEN_SIZE, 2, dropout=DROPOUT).to(DEVICE)
    opt = torch.optim.Adam(model.parameters(), lr=LR, weight_decay=1e-4)
    crit = nn.CrossEntropyLoss()

    model.train()
    for _ in range(EPOCHS):
        for xb, yb in dl_train:
            xb, yb = xb.to(DEVICE), yb.to(DEVICE)
            loss = crit(model(xb), yb)
            opt.zero_grad(); loss.backward(); opt.step()

    model.eval()
    ds_test = TensorDataset(torch.tensor(X_te_w, dtype=torch.float32))
    dl_test = DataLoader(ds_test, batch_size=BATCH_SIZE * 2, shuffle=False)

    all_preds = []
    with torch.no_grad():
        for (xb,) in dl_test:
            all_preds.append(torch.argmax(model(xb.to(DEVICE)), dim=1).cpu().numpy())
    flat_preds = np.concatenate(all_preds)
    preds_mat = flat_preds.reshape(len(y_test), n_times)
    return np.mean(preds_mat == y_test[:, None], axis=0)


def train_eegnet_windowed(X_train, y_train, X_test, y_test, seed=123):
    """Train EEGNet (windowed) and return per-time-point accuracy."""
    torch.manual_seed(seed)
    np.random.seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed_all(seed)
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False

    X_tr_w = make_sliding_windows_eegnet(X_train, WINDOW_SIZE)
    X_te_w = make_sliding_windows_eegnet(X_test, WINDOW_SIZE)
    n_times = X_train.shape[2]
    y_tr_flat = np.repeat(y_train[:, None], n_times, axis=1).flatten()

    ds_train = TensorDataset(torch.tensor(X_tr_w, dtype=torch.float32),
                             torch.tensor(y_tr_flat, dtype=torch.long))
    ds_test = TensorDataset(torch.tensor(X_te_w, dtype=torch.float32))
    dl_train = DataLoader(ds_train, batch_size=BATCH_SIZE, shuffle=True)
    dl_test = DataLoader(ds_test, batch_size=BATCH_SIZE * 2, shuffle=False)

    model = EEGNet_Small(n_channels=N_CHANNELS, n_classes=2, window_size=WINDOW_SIZE).to(DEVICE)
    opt = torch.optim.AdamW(model.parameters(), lr=EEGNET_LR, weight_decay=1e-3)
    crit = nn.CrossEntropyLoss()

    model.train()
    for _ in range(EEGNET_EPOCHS):
        for xb, yb in dl_train:
            xb, yb = xb.to(DEVICE), yb.to(DEVICE)
            loss = crit(model(xb), yb)
            opt.zero_grad(); loss.backward(); opt.step()

    model.eval()
    all_preds = []
    with torch.no_grad():
        for (xb,) in dl_test:
            all_preds.append(torch.argmax(model(xb.to(DEVICE)), dim=1).cpu().numpy())
    flat_preds = np.concatenate(all_preds)
    preds_mat = flat_preds.reshape(len(y_test), n_times)
    return np.mean(preds_mat == y_test[:, None], axis=0)


# ========================= 6. Main Benchmark Loop =========================

MODEL_NAMES = [
    "HeteroRC",
    "Std RNN", "Std LSTM", "Std Transformer",
    "Win RNN", "Win LSTM", "Win Transformer",
    "EEGNet",
]

results = {name: [] for name in MODEL_NAMES}
timing = {name: [] for name in MODEL_NAMES}
times = None

STD_REGISTRY = [
    ("Std RNN",         train_standard_model, StandardRNN),
    ("Std LSTM",        train_standard_model, StandardLSTM),
    ("Std Transformer", train_standard_model, StandardTransformer),
]
WIN_REGISTRY = [
    ("Win RNN",         train_windowed_model, WindowedRNN),
    ("Win LSTM",        train_windowed_model, WindowedLSTM),
    ("Win Transformer", train_windowed_model, WindowedTransformer),
]

print(f"\nStarting Benchmark (N={N_SUBJECTS}, {SIM_MODE} mode, hidden={HIDDEN_SIZE})...\n")

for sub in range(N_SUBJECTS):
    print(f"Subject {sub + 1}/{N_SUBJECTS} | ", end="", flush=True)

    X, y, t_vec = simulate_data(
        sub, mode=SIM_MODE, sfreq=SFREQ,
        n_trials=N_TOTAL_TRIALS, n_channels=N_CHANNELS,
    )
    if times is None:
        times = t_vec

    # --- HeteroRC ---
    print("RC.", end="", flush=True)
    t0 = time_module.time()
    acc_rc = time_resolved_decoding_heterorc(
        X, y, times=t_vec, n_folds=N_FOLDS,
        smooth_decisions=True, smooth_sigma_points=3.0,
        scale_percentile=ROBUST_PCTL, rc_params=rc_params, verbose=False,
    )
    timing["HeteroRC"].append(time_module.time() - t0)
    results["HeteroRC"].append(acc_rc)

    # --- CV splits (shared across all DL models) ---
    skf = StratifiedKFold(n_splits=N_FOLDS, shuffle=True, random_state=42)
    splits = list(skf.split(X, y))

    # --- Standard models ---
    for model_name, train_fn, model_cls in STD_REGISTRY:
        tag = model_name.split()[-1][0]  # R, L, T
        print(f"S{tag}.", end="", flush=True)
        t0 = time_module.time()
        fold_accs = []
        for tr, te in splits:
            scale = np.percentile(np.abs(X[tr]), ROBUST_PCTL) or 1.0
            fold_accs.append(train_fn(model_cls, X[tr] / scale, y[tr], X[te] / scale, y[te]))
        timing[model_name].append(time_module.time() - t0)
        results[model_name].append(np.mean(fold_accs, axis=0))

    # --- Windowed models ---
    for model_name, train_fn, model_cls in WIN_REGISTRY:
        tag = model_name.split()[-1][0]
        print(f"W{tag}.", end="", flush=True)
        t0 = time_module.time()
        fold_accs = []
        for tr, te in splits:
            scale = np.percentile(np.abs(X[tr]), ROBUST_PCTL) or 1.0
            fold_accs.append(train_fn(model_cls, X[tr] / scale, y[tr], X[te] / scale, y[te]))
        timing[model_name].append(time_module.time() - t0)
        results[model_name].append(np.mean(fold_accs, axis=0))

    # --- EEGNet ---
    print("E.", end="", flush=True)
    t0 = time_module.time()
    fold_accs = []
    for fold_idx, (tr, te) in enumerate(splits):
        scale = np.percentile(np.abs(X[tr]), ROBUST_PCTL) or 1.0
        fold_seed = 1000 * (sub + 1) + fold_idx
        fold_accs.append(train_eegnet_windowed(
            X[tr] / scale, y[tr], X[te] / scale, y[te], seed=fold_seed))
    timing["EEGNet"].append(time_module.time() - t0)
    results["EEGNet"].append(np.mean(fold_accs, axis=0))

    print(" Done.", flush=True)


# ========================= 7. Plotting =========================
plt.style.use('seaborn-v0_8-paper')
fig, axes = plt.subplots(1, 2, figsize=(22, 8), sharey=True)

COLORS = {
    "HeteroRC":        "#e74c3c",
    "Std RNN":         "#e67e22",
    "Std LSTM":        "#3498db",
    "Std Transformer": "#9b59b6",
    "Win RNN":         "#e67e22",
    "Win LSTM":        "#3498db",
    "Win Transformer": "#9b59b6",
    "EEGNet":          "#2ecc71",
}

LABELS = {
    "HeteroRC":        "HeteroRC",
    "Std RNN":         "Standard RNN (32)",
    "Std LSTM":        "Standard LSTM (32)",
    "Std Transformer": "Standard Transformer (32)",
    "Win RNN":         "Windowed RNN (32)",
    "Win LSTM":        "Windowed LSTM (32)",
    "Win Transformer": "Windowed Transformer (32)",
    "EEGNet":          "EEGNet",
}


def plot_panel(ax, model_keys, title):
    for name in model_keys:
        data = np.array(results[name])
        mean = np.mean(data, axis=0)
        se = sem(data, axis=0)
        if name != "HeteroRC":
            mean = gaussian_filter1d(mean, sigma=2.0)
        lw = 3.0 if name == "HeteroRC" else 2.0
        ls = "-"
        if name.startswith("Std "):
            ls = "--"
        ax.plot(times, mean, label=LABELS[name], color=COLORS[name],
                linestyle=ls, linewidth=lw)
        ax.fill_between(times, mean - se, mean + se, color=COLORS[name], alpha=0.1)
    ax.axhline(0.5, linestyle=':', color='gray')
    ax.axvspan(0.2, 0.6, color='gray', alpha=0.05)
    ax.set_ylim(0.35, 1.0)
    ax.set_xlabel('Time (s)', fontsize=14)
    ax.set_title(title, fontsize=16)
    ax.legend(fontsize=10, loc='upper left', frameon=True)


# Left panel: Standard models
plot_panel(axes[0],
           ["HeteroRC", "Std RNN", "Std LSTM", "Std Transformer", "EEGNet"],
           "Standard (Seq-to-Seq) Models")
axes[0].set_ylabel('Decoding Accuracy', fontsize=14)

# Right panel: Windowed models
plot_panel(axes[1],
           ["HeteroRC", "Win RNN", "Win LSTM", "Win Transformer", "EEGNet"],
           "Windowed (Seq-to-One) Models")

fig.suptitle(
    f'Model Comparison ({SIM_MODE} mode, N={N_SUBJECTS}, hidden={HIDDEN_SIZE})',
    fontsize=18, y=1.02,
)

plt.tight_layout()
save_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                         "Transformer_RNN_LSTM_comparison.png")
plt.savefig(save_path, dpi=150, bbox_inches='tight')
print(f"\nFigure saved to: {save_path}")
plt.close()


# ========================= 8. Save to Excel =========================
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                          "Model_Comparison_Results.xlsx")

wb = openpyxl.Workbook()

header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)


def style_header(ws, row, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row, c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


# --- Per-model accuracy sheets (subject x time) ---
for name in MODEL_NAMES:
    ws = wb.create_sheet(title=name[:31])
    ws.cell(1, 1, "Subject")
    for j, t in enumerate(times):
        ws.cell(1, j + 2, round(float(t), 3))
    style_header(ws, 1, len(times) + 1)
    ws.column_dimensions['A'].width = 12

    data = results[name]
    for i, row in enumerate(data):
        ws.cell(i + 2, 1, f"Sub {i + 1}")
        ws.cell(i + 2, 1).border = thin_border
        for j, val in enumerate(row):
            cell = ws.cell(i + 2, j + 2, round(float(val), 4))
            cell.border = thin_border
            cell.number_format = '0.0000'

    # Mean and SEM rows
    r_mean = len(data) + 2
    r_sem = len(data) + 3
    ws.cell(r_mean, 1, "Mean")
    ws.cell(r_sem, 1, "SEM")
    ws.cell(r_mean, 1).font = Font(bold=True)
    ws.cell(r_sem, 1).font = Font(bold=True)
    ws.cell(r_mean, 1).border = thin_border
    ws.cell(r_sem, 1).border = thin_border
    arr = np.array(data)
    for j in range(arr.shape[1]):
        cell_m = ws.cell(r_mean, j + 2, round(float(np.mean(arr[:, j])), 4))
        cell_s = ws.cell(r_sem, j + 2, round(float(sem(arr[:, j])), 4))
        cell_m.number_format = '0.0000'
        cell_s.number_format = '0.0000'
        cell_m.font = Font(bold=True)
        cell_m.border = thin_border
        cell_s.border = thin_border

# --- Timing sheet ---
ws_t = wb.create_sheet(title="Timing (s per subject)")
ws_t.cell(1, 1, "Subject")
for j, name in enumerate(MODEL_NAMES):
    ws_t.cell(1, j + 2, name)
style_header(ws_t, 1, len(MODEL_NAMES) + 1)
ws_t.column_dimensions['A'].width = 12

for i in range(N_SUBJECTS):
    ws_t.cell(i + 2, 1, f"Sub {i + 1}")
    ws_t.cell(i + 2, 1).border = thin_border
    for j, name in enumerate(MODEL_NAMES):
        cell = ws_t.cell(i + 2, j + 2, round(timing[name][i], 2))
        cell.number_format = '0.00'
        cell.border = thin_border

# Mean timing row
r_mean = N_SUBJECTS + 2
ws_t.cell(r_mean, 1, "Mean")
ws_t.cell(r_mean, 1).font = Font(bold=True)
ws_t.cell(r_mean, 1).border = thin_border
for j, name in enumerate(MODEL_NAMES):
    cell = ws_t.cell(r_mean, j + 2, round(float(np.mean(timing[name])), 2))
    cell.font = Font(bold=True)
    cell.number_format = '0.00'
    cell.border = thin_border

# Auto-width for timing columns
for j in range(2, len(MODEL_NAMES) + 2):
    ws_t.column_dimensions[openpyxl.utils.get_column_letter(j)].width = 18


# --- Peak Accuracy Summary sheet ---
ws_p = wb.create_sheet(title="Peak Summary")
p_headers = ["Model", "Peak Accuracy", "Peak Time (s)", "Beat HeteroRC?",
             "Gap vs RC", "Avg Time/Subject (s)"]
for j, h in enumerate(p_headers):
    ws_p.cell(1, j + 1, h)
style_header(ws_p, 1, len(p_headers))

# RC peak for comparison
rc_mean = np.mean(np.array(results["HeteroRC"]), axis=0)
rc_peak = np.max(rc_mean)

yes_font = Font(color='006100', bold=True)
no_font = Font(color='9C0006')
rc_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

for i, name in enumerate(MODEL_NAMES):
    data = np.array(results[name])
    mean = np.mean(data, axis=0)
    if name != "HeteroRC":
        mean = gaussian_filter1d(mean, sigma=2.0)
    peak_idx = np.argmax(mean)
    peak_acc = mean[peak_idx]
    peak_time = times[peak_idx]
    avg_time = np.mean(timing[name])

    if name == "HeteroRC":
        beat_str = "—"
        gap_val = "—"
    else:
        beat_str = "YES" if peak_acc > rc_peak else "no"
        gap_val = round(float(peak_acc - rc_peak), 4)

    row = i + 2
    ws_p.cell(row, 1, name)
    ws_p.cell(row, 2, round(float(peak_acc), 4))
    ws_p.cell(row, 2).number_format = '0.0000'
    ws_p.cell(row, 3, round(float(peak_time), 3))
    ws_p.cell(row, 4, str(beat_str))
    ws_p.cell(row, 5, gap_val if isinstance(gap_val, str) else gap_val)
    if isinstance(gap_val, float):
        ws_p.cell(row, 5).number_format = '0.0000'
    ws_p.cell(row, 6, round(float(avg_time), 2))
    ws_p.cell(row, 6).number_format = '0.00'

    for c in range(1, len(p_headers) + 1):
        ws_p.cell(row, c).border = thin_border

    if name == "HeteroRC":
        for c in range(1, len(p_headers) + 1):
            ws_p.cell(row, c).fill = rc_fill
    elif beat_str == "YES":
        ws_p.cell(row, 4).font = yes_font
    else:
        ws_p.cell(row, 4).font = no_font

ws_p.cell(len(MODEL_NAMES) + 3, 1,
          f"HeteroRC peak accuracy: {rc_peak:.4f}")
ws_p.cell(len(MODEL_NAMES) + 3, 1).font = Font(italic=True, color='666666')

for j, w in enumerate([22, 15, 15, 15, 12, 20]):
    ws_p.column_dimensions[openpyxl.utils.get_column_letter(j + 1)].width = w

# Remove default sheet
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

wb.save(EXCEL_PATH)
print(f"Excel saved to: {EXCEL_PATH}")


# ========================= 9. Console Summary =========================
print("\n" + "=" * 95)
print("PEAK ACCURACY SUMMARY (mean across subjects)")
print("=" * 95)
print(f"{'Model':<25} {'Peak Acc':>10} {'Peak Time (s)':>14} {'Beat RC?':>10} "
      f"{'Gap vs RC':>10} {'Avg Time (s)':>14}")
print("-" * 95)

for name in MODEL_NAMES:
    data = np.array(results[name])
    mean = np.mean(data, axis=0)
    if name != "HeteroRC":
        mean = gaussian_filter1d(mean, sigma=2.0)
    peak_idx = np.argmax(mean)
    peak_acc = mean[peak_idx]
    peak_time = times[peak_idx]
    avg_time = np.mean(timing[name])

    if name == "HeteroRC":
        print(f"{name:<25} {peak_acc:>10.3f} {peak_time:>14.3f} "
              f"{'---':>10} {'---':>10} {avg_time:>14.2f}")
    else:
        beat = "YES" if peak_acc > rc_peak else "no"
        gap = peak_acc - rc_peak
        print(f"{name:<25} {peak_acc:>10.3f} {peak_time:>14.3f} "
              f"{beat:>10} {gap:>+10.3f} {avg_time:>14.2f}")

print("=" * 95)

print("\nTIMING SUMMARY (seconds per subject, mean +/- std)")
print("-" * 60)
for name in MODEL_NAMES:
    avg_t = np.mean(timing[name])
    std_t = np.std(timing[name])
    print(f"  {name:<25} {avg_t:>8.2f} +/- {std_t:.2f}")
print("-" * 60)
